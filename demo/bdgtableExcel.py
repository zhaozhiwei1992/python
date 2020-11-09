"""
# 将指标系统表转换为多个sheet信息
# sheet为表名
# 每个sheet为对应表信息

# 读取oracle表信息

#  更新列的中文字段信息
update BDG_TABLE_INFO t set t.COMENT1 = (select name from PAY_LHC170119.FASP_T_DICCOLUMN t2 where t2.tablecode = 'BDG_T_BDGMAIN' and t2.columncode = t.COLUMN_NAME) where t.COMENT1 is null;

"""


import cx_Oracle
import os

# 设置查询编码
os.environ['NLS_LANG'] = 'AMERICAN_AMERICA.ZHS16GBK'

def getBdgTableInfo():

    tableinfos = [];

    #字段信息
    con = cx_Oracle.connect('pay_lhc170119/1@192.168.3.6/orcl')
    cur = con.cursor()
    sql = "select t.TABLE_NAME, t.COLUMN_NAME, t.COMENT1 , t.DATA_TYPE, t.DATA_LENGTH from bdg_table_info t order by t.TABLE_NAME"
    cur.prepare(sql)
    cur.execute(None)
    res = cur.fetchall()
    index = len(res)
    for result in res:
        index -= 1

        tableInfo = {}
        tableInfo['table_name'] = result[0]
        tableInfo['column_ename'] = result[1]
        tableInfo['column_cname'] = result[2]
        tableInfo['data_type'] = result[3]
        tableInfo['data_length'] = result[4]
        tableInfo['isnull'] = ""
        tableInfo['default'] = ""
        tableInfo['remark'] = ""
        tableinfos.append(tableInfo);
    return tableinfos

from operator import itemgetter #itemgetter用来去dict中的key，省去了使用lambda函数
from itertools import groupby #itertool还包含有其他很多函数，比如将多个list联合起来。。

from openpyxl import Workbook


# 根据table_name分组
# 表明作为sheet信息
# value作为列信息
def createExcel(datas):
    #  创建excel
    wb = Workbook()
    wb.active

    for key, group in datas:
        #  创建标签
        ws1 = wb.create_sheet(key) # 新建sheet,插入到最后(默认)

#  字段名 字段中文名 类型 长度 是否可为空 默认值 备注
        # 创建表头
        ws1.cell(1, 1, "字段名 ")
        ws1.cell(1, 2, "字段中文名 ")
        ws1.cell(1, 3, "类型")
        ws1.cell(1, 4, "长度 ")
        ws1.cell(1, 5, "是否可为空")
        ws1.cell(1, 6, "默认值 ")
        ws1.cell(1, 7, "备注")
        #  填入数据
        i = 1
        for g in group:
            count = 0
            i=i+1
            for (k,v) in  g.items(): 
                # tablename信息在没行中直接丢弃
                if(k == 'table_name' ): continue
                count+=1
                # 地一个跳过
                # print "dict[%s]=" % k,v 
                ws1.cell(i, count, v)

    # 写出
    wb.save("/tmp/bdgtableinfos.xlsx")

if __name__ == '__main__':
    infos = getBdgTableInfo()
    # pprint(infos)
    print(infos)

    # 数据根据table_name分组
    infosGroupByTableName = groupby(infos,itemgetter('table_name')) 
    # for key,group in lstg:
    #     for g in group: #group是一个迭代器，包含了所有的分组列表
    #         print (key,g)

    createExcel(infosGroupByTableName);
