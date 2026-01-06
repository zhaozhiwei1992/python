"""

# 将指标系统表转换为多个sheet信息
# sheet为表名
# 每个sheet为对应表信息

"""
import openpyxl
from openpyxl import Workbook



# 根据table_name分组
# 表明作为sheet信息
# value作为列信息
def createExcel(datas):
    """
    优化后的Excel创建函数，支持大数据量写入
    """
    # 使用write_only模式提高写入性能[5,6](@ref)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="项目信息")

    # 添加表头（根据实际需求调整）
    header = ["合并后的项目信息"]
    ws.append(header)

    # 批量写入数据
    for data in datas:
        ws.append([data])

    wb.save("/tmp/项目信息合并.xlsx")


if __name__ == '__main__':

    # 1. 提取并合并列信息
    print('Opening workbook...')
    wb = openpyxl.load_workbook('/home/zhaozhiwei/Documents/2026年省级项目支出预算标准表.xlsx', read_only=True)
    print("所有sheet", wb.sheetnames)
    # 获取sheet
    # sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    sheet = wb[wb.sheetnames[0]]
    # Fill in countyData with each county's population and tracts.
    print('Reading rows...')
    # 从第一行开始，第一行一般为标题行
    result = []
    # 性能贼差的凡是
    # for row in range(1, len(tuple(sheet.rows))):
    #     # Each row in the spreadsheet has data for one census tract.
    #     # B:项目编码 和 G:项目内容合并
    #     try:
    #         content = '项目编码: \n' + str(sheet['B' + str(row)].value)
    #         content += '\n 项目内容: \n' + str(sheet['G' + str(row)].value)
    #     except:
    #         print(row + " 行报错了")
    #
    #     result.append(content)

    # 性能很好，可以采用
    for row_idx, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), 1):
        try:
            if len(row) >= 12:  # 确保有足够的列
                d_content = row[3] or ""
                g_content = row[6] or ""
                m_content = row[12] or ""

                content = f"\n{d_content}-{g_content}\n{m_content}\n"
                result.append(content)

            # 分批处理，避免内存溢出[6](@ref)
            if len(result) >= 10000:  # 每10000行处理一次
                print(f"已处理 {row_idx} 行...")
                # 这里可以添加分批保存逻辑

        except Exception as e:
            print(f"第 {row_idx} 行处理出错: {str(e)}")
            continue

    wb.close()

    # print('#####\n'.join(result))
    # 3. 生成新的excel
    # createExcel(result)

    # 4. 写出到txt文件
    with open("/tmp/项目信息合并.txt", "w") as f:
        f.write('#####\n'.join(result))
