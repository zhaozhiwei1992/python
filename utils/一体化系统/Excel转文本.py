from openpyxl import load_workbook

# 加载Excel文件
wb = load_workbook("/home/zhaozhiwei/vagrant/win10/input.xlsx")
sheet = wb.active

# 提取表头
headers = [cell.value for cell in sheet[1]]

# 逐行处理并写入文本文件
with open("/home/zhaozhiwei/vagrant/win10/output.txt", "w", encoding="utf-8") as f:
    for row in sheet.iter_rows(min_row=2):  # 跳过标题行
        fields = []
        for idx, cell in enumerate(row):
            value = str(cell.value).strip() if cell.value else ""
            if value:
                fields.append(f"{headers[idx]}:{value}")
        f.write(", ".join(fields) + "\n")