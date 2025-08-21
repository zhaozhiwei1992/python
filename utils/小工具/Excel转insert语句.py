import openpyxl

wb = openpyxl.load_workbook('/tmp/area.xlsx', False)

# sheet = wb.get_sheet_by_name(wb.sheetnames[0])
sheet = wb[wb.sheetnames[0]]
# Fill in countyData with each county's population and tracts.
print('Reading rows...')

headers = sheet.rows[0]
headerStr = ', '.join(headers)

insert_statements = []
# 从第一行开始，第一行一般为标题行
for row in range(0, len(tuple(sheet.rows))):
    if row:  # Ensure it's not an empty row
        values = ["'" + value + "'" if value else "''" for value in row]
        formatted_values = ', '.join(values)
        insert_statement = f"INSERT INTO sys_area ({headerStr}) VALUES ({formatted_values});"
        insert_statements.append(insert_statement)

# Now you have a list of all INSERT statements
for statement in insert_statements:
    print(statement)