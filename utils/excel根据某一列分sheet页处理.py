"""

# 将指标系统表转换为多个sheet信息
# sheet为表名
# 每个sheet为对应表信息

"""

# itemgetter用来去dict中的key，省去了使用lambda函数
from operator import itemgetter
# itertool还包含有其他很多函数，比如将多个list联合起来。。
from itertools import groupby

import openpyxl
from openpyxl import Workbook


def getColsInfoFromExcel(cols):
    tableinfos = []
    print('Opening workbook...')
    wb = openpyxl.load_workbook('/mnt/d/vagrant/win7/指标下达.xlsx')
    print("所有sheet", wb.sheetnames)
    # 获取sheet
    # sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    sheet = wb[wb.sheetnames[0]]
    # Fill in countyData with each county's population and tracts.
    print('Reading rows...')
    # 从第一行开始，第一行一般为标题行
    for row in range(1, len(tuple(sheet.rows))):
        # Each row in the spreadsheet has data for one census tract.
        # 从B开始, 第一行数据为标题, 保留每行信息
        # sheet['I' + str(row)].value
        tableInfo = {}
        for v in cols:
            tableInfo[v] = sheet[v + str(row)].value
        tableinfos.append(tableInfo)
    return tableinfos


# 根据table_name分组
# 表明作为sheet信息
# value作为列信息
def createExcel(datas):
    #  创建excel
    wb = Workbook()
    wb.active

    titleInfo = {}
    titleFlag = True
    for key, group in datas:
        #  创建标签
        # 新建sheet,插入到最后(默认), 标题列不能创建, 默认为第一行
        # data为分组信息
        if titleFlag:
            titleFlag = False
            for g in group:
                titleInfo = g
            pass
        else:
            ws1 = wb.create_sheet(key)
            # 创建表头
            for index, v in enumerate(titleInfo):
                ws1.cell(1, index + 1, titleInfo.get(v))

        #  填入数据
        i = 1
        # row, map list [{}, {}...]
        for g in group:
            count = 0
            i += 1
            # cols  {"B":"B_value", "C":"C_value", ...}
            for (k, v) in g.items():
                # 分组列信息直接丢弃, 只保留后边的
                # if k == 'B':
                #     continue
                count += 1
                # 第一个跳过
                # print "dict[%s]=" % k,v 
                ws1.cell(i, count, v)

    # 写出
    wb.save("/tmp/指标下达_sheet.xlsx")


if __name__ == '__main__':

    # 列标集合
    cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
    infos = getColsInfoFromExcel(cols)
    # print(infos)

    # 数据根据第一列分组
    infosGroupByTableName = groupby(infos, itemgetter(cols[0]))

    createExcel(infosGroupByTableName)
