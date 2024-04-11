"""
1. diff: 保留原始对比结果(SRC_FILE_PATH), 写入单独文档,每次发布新的明细报告先和原始文档比对一次，防止重复工作
2. merge: 处理结果合并到SRC_FILE_PATH对应的文件中

使用
    上报数据完整性明细报告.py diff
    上报数据完整性明细报告.py merge
"""
import sys

import openpyxl
from pathlib import Path

# 文件路径常量
SRC_FILE_PATH = Path("/d/at_work/一体化系统文档/规范/上报数据完整性明细报告-浙江支付.xlsx")
DEST_FILE_PATH = Path("/home/zhaozhiwei/vagrant/win7/上报数据完整性明细报告20240409.xlsx")


def read_workbook(file_path):
    """读取工作簿"""
    return openpyxl.load_workbook(file_path, read_only=False)


def save_workbook(workbook, file_path):
    """保存工作簿"""
    workbook.save(file_path)
    workbook.close()


def fill_data(src_sheet, dest_sheet):
    """
    填充数据
    根据指定单元格内容相同则进行数据填充
    """
    # 遍历原始文档，填充数据到目标文档
    for dest_row in dest_sheet.iter_rows(min_row=3):
        dest_value = dest_row[6].value  # 获取目标表格中当前行第6列的值
        if dest_value:  # 如果第6列有值
            for src_row in src_sheet.iter_rows(min_row=3):
                if src_row[6].value == dest_value and src_row[1].value:  # 匹配第6列的值并且源表格第1列有值
                    dest_row[1].value = src_row[1].value  # 覆盖目标表格第1列的值
                    break  # 匹配到一次即可跳出内层循环


def diff():
    """对比原始文档和目标文档，填充数据"""
    # 读取原始文档和目标文档
    src_wb = read_workbook(SRC_FILE_PATH)
    dest_wb = read_workbook(DEST_FILE_PATH)

    # 处理每个工作表
    for sheet_name in ['完整性-表内勾稽关系', '完整性-表间勾稽关系']:
        src_sheet = src_wb[sheet_name]
        dest_sheet = dest_wb[sheet_name]
        fill_data(src_sheet, dest_sheet)

    # 保存目标文档
    save_workbook(dest_wb, DEST_FILE_PATH)


def merge():
    """
    合并其他操作
    将DIST_FILE_PATH对应的内容,筛选后加入到SRC_FILE_PATH中,只保留不重复部分
    """
    """合并数据"""
    src_wb = read_workbook(SRC_FILE_PATH)
    dest_wb = read_workbook(DEST_FILE_PATH)

    for sheet_name in ['完整性-表内勾稽关系', '完整性-表间勾稽关系']:
        # 获取源和目标表格
        src_sheet = src_wb[sheet_name]
        dest_sheet = dest_wb[sheet_name]
        # 构建源表格中已有数据的集合，以加快查找速度
        src_set = set()
        for row in src_sheet.iter_rows(min_row=3, values_only=True):
            if row[6]:  # 只关心第6列的内容
                src_set.add(row[6])

        # 遍历目标表格，将不存在于源表格中的数据写入源表格
        for row in dest_sheet.iter_rows(min_row=3, values_only=True):
            if row[6] not in src_set and ("PAY_" in str(row[2]) or "INC_" in str(row[2])):  # 第2列有值且第6列的值不在源表格中
                src_sheet.append(row)

    # 保存源工作簿
    save_workbook(src_wb, SRC_FILE_PATH)

# 注册动态方法
dynamic_func = {
    "diff": diff,
    "merge": merge,
}


def main():
    op = sys.argv[1]
    if op:
        dynamic_func[op]()
    else:
        # 默认比较不同,并填充描述
        diff()


if __name__ == '__main__':
    main()
