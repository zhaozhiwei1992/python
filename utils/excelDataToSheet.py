"""

# 将指标系统表转换为多个sheet信息
# sheet为表名
# 每个sheet为对应表信息

"""

# itemgetter用来去dict中的key，省去了使用lambda函数
from operator import itemgetter
# itertool还包含有其他很多函数，比如将多个list联合起来。。
from itertools import groupby

import openpyxl
from openpyxl import Workbook

def getColsInfoFromExcel():
    tableinfos = []
    print('Opening workbook...')
    wb = openpyxl.load_workbook('/mnt/d/vagrant/win7/指标下达.xlsx')
    print("所有sheet", wb.sheetnames)
    # 获取sheet
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    # Fill in countyData with each county's population and tracts.
    print('Reading rows...')
    for row in range(2, len(tuple(sheet.rows))):
        # Each row in the spreadsheet has data for one census tract.
        # A区划编码
        # B区划名称
        # C省级文号
        # D省级项目
        # E下达金额
        # F预算形式
        # G转移支付功能科目
        # H支出功能科目
        # 从B开始
        # sheet['I' + str(row)].value
        tableInfo = {}
        tableInfo['B'] = sheet['B' + str(row)].value
        tableInfo['C'] = sheet['C' + str(row)].value
        tableInfo['D'] = sheet['D' + str(row)].value
        tableInfo['E'] = sheet['E' + str(row)].value
        tableInfo['F'] = sheet['F' + str(row)].value
        tableInfo['G'] = sheet['G' + str(row)].value
        tableInfo['H'] = sheet['H' + str(row)].value
        tableinfos.append(tableInfo)
    return tableinfos


# 根据table_name分组
# 表明作为sheet信息
# value作为列信息
def createExcel(datas):
    #  创建excel
    wb = Workbook()
    wb.active

    for key, group in datas:
        #  创建标签
        ws1 = wb.create_sheet(key)  # 新建sheet,插入到最后(默认)

        #  字段名 字段中文名 类型 长度 是否可为空 默认值 备注
        # 创建表头
        ws1.cell(1, 1, "省级文号")
        ws1.cell(1, 2, "省级项目 ")
        ws1.cell(1, 3, "下达金额")
        ws1.cell(1, 4, "预算形式")
        ws1.cell(1, 5, "转移支付功能科目")
        ws1.cell(1, 6, "支出功能科目")
        #  填入数据
        i = 1
        for g in group:
            count = 0
            i = i + 1
            for (k, v) in g.items():
                # 分组信息信息直接丢弃
                if k == 'B':
                    continue
                count += 1
                # 地一个跳过
                # print "dict[%s]=" % k,v 
                ws1.cell(i, count, v)

    # 写出
    wb.save("/tmp/指标下达_sheet.xlsx")


if __name__ == '__main__':
    infos = getColsInfoFromExcel()
    # pprint(infos)
    print(infos)

    # 数据根据指定列 B 分组
    infosGroupByTableName = groupby(infos, itemgetter('B'))

    createExcel(infosGroupByTableName)
