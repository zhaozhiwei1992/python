"""
每个版本执行一次, 初始化

传入版本参数

python /home/zhaozhiwei/workspace/python/utils/产品发布说明初始化.py PAYZJ VERSION

"""
import datetime
import sys
import openpyxl
from pathlib import Path

# 直接从 fborg 模块导入函数
from 项目.获取发版内容org import find as find_in_org


def main():
    # 获取当前日期
    now = datetime.date.today()
    print(now)

    # 直接提取年份和月份
    year, month, day = now.year, now.month, now.day

    # 读取命令行参数并检查参数数量
    if len(sys.argv) < 3:
        print("用法: python script.py <APP_ID> <VERSION>")
        return

    appid, version = sys.argv[1], sys.argv[2]

    # 使用 Pathlib 构建文件路径，更加清晰和高效
    if appid.upper() in ("NFCS", "ISA"):
        fbFile = Path(f"/home/zhaozhiwei/workspace/{appid.upper()}/发布说明/版本发布计划.xlsx")
    else:
        fbFile = Path(f"/home/zhaozhiwei/workspace/ifmis-4.1/ifmis-{appid.lower()}/发布说明/版本发布计划.xlsx")

    # 使用 openpyxl 加载工作簿
    wb = openpyxl.load_workbook(fbFile, read_only=False)
    sheet = wb['版本发布计划']
    appName = sheet['C4'].value

    print('读取版本发布计划的行...')
    allRows = []
    for row in sheet.iter_rows(min_row=9, values_only=True):
        rowData = {'content': row[2], 'srcId': row[3]}
        allRows.append(rowData)

    wb.close()

    # 调整发布说明内容
    if appid.upper() in ("NFCS", "ISA"):
        fbFile = Path(f"/home/zhaozhiwei/workspace/{appid.upper()}/发布说明/产品发布说明模板.xlsx")
    else:
        fbFile = Path(f"/home/zhaozhiwei/workspace/ifmis-4.1/ifmis-{appid.lower()}/发布说明/产品发布说明模板.xlsx")

    wb = openpyxl.load_workbook(fbFile, read_only=False)
    sheet = wb['01_发版说明']

    # 修改版本信息
    sheet['F2'].value = appid
    sheet['C2'].value = appName
    sheet['H2'].value = f"{month}月{day}日"
    sheet['H15'].value = f"{month}月{day}日"
    sheet['B7'].value = f"{appName}后端"
    sheet['B8'].value = f"{appName}前端"
    sheet['C7'].value = version.replace("_", ".")
    sheet['C8'].value = version.replace("_", ".")

    if appid.upper() == "NFCS":
        sheet['E7'].value = f"ifmis-data-center-{version.replace('_', '.')}-SNAPSHOT.jar"
        sheet['E8'].value = ""
    elif appid.upper() == "ISA":
        sheet['E7'].value = f"ifmis-service-agent-{version.replace('_', '.')}-SNAPSHOT.jar"
        sheet['E8'].value = ""
    else:
        sheet['E7'].value = f"ifmis-{appid.lower()}-service-{version.replace('_', '.')}-SNAPSHOT.jar"
        sheet['E8'].value = f"ifmis-{appid.lower()}-webapp-{version.replace('_', '.')}-SNAPSHOT.jar"

    sheet['G7'].value = version.replace("_", ".")
    sheet['G8'].value = version.replace("_", ".")

    # 读取依赖信息
    fbContentMap = find_in_org(appid.upper(), f"V_{version}")
    deptList = fbContentMap.get("dept")

    for rowIndex, deptTuple in enumerate(deptList, start=12):
        sheet[f'B{rowIndex}'].value = deptTuple[0]
        sheet[f'E{rowIndex}'].value = deptTuple[1]

    # 更新 '02_发版测试' 表
    sheet = wb['02_发版测试']
    sheet['F2'].value = f"{month}月{day}日"
    sheet['E3'].value = f"{year}/{month - 1}/{day}-{year}/{month}/{day}"

    # 更新 '03_版本更新内容' 表
    sheet = wb['03_版本更新内容']
    for rowIndex, rowData in enumerate(allRows, start=3):
        sheet.cell(rowIndex, 1, rowIndex - 2)
        srcId = rowData['srcId']
        sheet.cell(rowIndex, 2, '问题' if srcId and '#' in str(srcId) else '需求')
        sheet.cell(rowIndex, 3, rowIndex - 2)
        sheet.cell(rowIndex, 4, '预算执行')
        sheet.cell(rowIndex, 5, appName)
        sheet.cell(rowIndex, 6, rowData['content'])
        sheet.cell(rowIndex, 7, rowData['srcId'])

    # 保存工作簿
    wb.save(fbFile)


if __name__ == '__main__':
    main()
