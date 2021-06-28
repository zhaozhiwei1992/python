#! python3
"""
统计当前问题列表, 并发邮件记录
python QuestionTODO.py 文件名 --date 日期(2020/03/04)
"""
# 读取制定日期行
# 输出文本，或发邮件
# readCensusExcel.py - Tabulates population and number of census tracts for
# each county.
# import pdb
import openpyxl, pprint
import datetime
# import questtionDatas
# print(questtionDatas.allData)

print('Opening workbook...')
wb = openpyxl.load_workbook('/home/zhaozhiwei/workspace/python/utils/惠民系统问题及需求更新（含用例）(1).xlsx')
print("所有sheet", wb.sheetnames)
# 获取sheet
sheet = wb.get_sheet_by_name(wb.sheetnames[2])
# pdb.set_trace()
questionDatas = {}
# Fill in countyData with each county's population and tracts.
print('Reading rows...')
for row in range(3, len(tuple(sheet.rows))):
    # Each row in the spreadsheet has data for one census tract.
    # 获取当前行日期
    date = sheet['I' + str(row)].value
    date = datetime.datetime.strftime(date, "%Y%m%d")

    questionMsg = sheet['G' + str(row)].value

    # 保存为日期键值对
    # Make sure the key for this state exists.
    questionDatas.setdefault(date, [])
    questionDatas.get(date).append(questionMsg)

print(questionDatas)

# # Open a new text file and write the contents of countyData to it.
print('Writing results...')
resultFile = open('/tmp/questtionDatas.py', 'w')
resultFile.write('allData = ' + pprint.pformat(questionDatas))
resultFile.close()
print('Done.')
